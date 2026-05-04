from DB import DataBase
from openpyxl import load_workbook

students = load_workbook('Data/students.xlsx').active
type_of_learning = load_workbook('Data/type_of_learning.xlsx').active
groups = load_workbook('Data/groups.xlsx').active
levels = load_workbook('Data/levels.xlsx').active

db = DataBase()
db.executemany(
    "INSERT OR REPLACE INTO `types_learning` (`id_type_learning`, `type_name`) VALUES (?, ?)",
    [(t[0], t[1]) for t in type_of_learning.iter_rows(min_row=2, values_only=True)],
)
db.executemany(
    "INSERT OR REPLACE INTO `groups` (`id_group`, `group_name`) VALUES (?, ?)",
    [(g[0], g[1]) for g in groups.iter_rows(min_row=2, values_only=True)],
)
db.executemany(
    "INSERT OR REPLACE INTO `levels` (`id_level`, `level_name`) VALUES (?, ?)",
    [(l[0], l[1]) for l in levels.iter_rows(min_row=2, values_only=True)],
)
db.executemany(
    "INSERT OR REPLACE INTO `students` (`id_level`, `id_group`, `id_type_learning`, `name`, `surname`, `father_name`, `age`, `grade_mean`) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",    
    [(s[0], s[1], s[2], s[3], s[4], s[5], s[6], s[7]) for s in students.iter_rows(min_row=2, values_only=True)],
)                    



count_students = db.query('''
SELECT id_student FROM students;
''')
print("Кол-во студентов:", len(count_students))

count_students_by_type = db.query('''
SELECT t.group_name, COUNT(s.id_student) AS student_count
FROM students s
JOIN groups t ON s.id_group = t.id_group
GROUP BY t.group_name;
                                                                                   '''
)

print("Кол-во студентов по группам:")
for group_name, student_count in count_students_by_type:
    print(f"{group_name}: {student_count}")


count_students_by_type =db.query('''
SELECT t.type_name, COUNT(s.id_student) AS student_count
FROM students s
JOIN types_learning t ON s.id_type_learning = t.id_type_learning
                                 GROUP BY t.type_name;''')

print("Кол-во студентов по типу обучения:")
for type_name, student_count in count_students_by_type:
    print(f"{type_name}: {student_count}")  

mark_max_min = db.query('''
SELECT MAX(grade_mean) AS max_grade, MIN(grade_mean) AS min_grade, AVG(grade_mean) AS avg_grade FROM students;
''')
print(f"Макс. средняя оценка: {mark_max_min [0][0]}, Мин. средняя оценка: {mark_max_min [0][1]}, Средняя оценка: {mark_max_min [0][2]}")  

mark_by_type_and_group_and_level = db.query('''
SELECT t.type_name, g.group_name, l.level_name, AVG(s.grade_mean) AS avg_grade
FROM students s
JOIN types_learning t ON s.id_type_learning = t.id_type_learning
JOIN groups g ON s.id_group = g.id_group
JOIN levels l ON s.id_level = l.id_level
GROUP BY t.type_name, g.group_name, l.level_name;
''')
print(mark_by_type_and_group_and_level)
print("Средняя оценка по типу обучения, группе и уровню:")
for type_name, group_name, level_name, avg_grade in mark_by_type_and_group_and_level:
    print(f"Тип обучения: {type_name}, Группа: {group_name}, Уровень: {level_name}, Средняя оценка: {avg_grade}")


db.closeDB()
# mark_max_min [0][1]=np.asarray(mark_max_min [0][1], dtype=np.float64)
# mark_max_min [0][0]=np.asarray(mark_max_min [0][0], dtype=np.float64)
# print(f"Макс. средняя оценка: {mark_max_min [0][0]}, Мин. средняя оценка: {mark_max_min [0][1]}")